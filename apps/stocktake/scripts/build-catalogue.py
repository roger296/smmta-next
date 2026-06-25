#!/usr/bin/env python3
"""
Build the stock-take-lite seed catalogue from the Big Bakes stock-take
spreadsheet.

The June 2026 sheet is the canonical blank template the managers fill in. Its
structure is encoded in cell formatting, not data: a **bold** cell in column A
is a heading, a non-bold cell is a countable item. A bold row immediately
followed by another bold row is a top-level *area* (e.g. "Bar Stock"); the
second bold row is the *section* under it (e.g. "Wines"). Column B is the
pack-size hint ("25kg", "1 x 70cl"), column C the usual supplier.

SheetJS (the repo's Node xlsx lib) can't read the bold flag — that's why this
generator is Python+openpyxl. It's a one-off seed step; re-run it only when the
master spreadsheet's item list changes.

Usage:
  python apps/stocktake/scripts/build-catalogue.py \
      ["C:\\path\\to\\Stocktake.xlsx"] ["Stockcount List JUNE 2026"]

Defaults point at the source spreadsheet on Roger's machine. Writes
apps/stocktake/src/data/catalogue.json (relative to the repo root).
"""
import sys, os, re, json, unicodedata

import openpyxl

DEFAULT_SRC = r"C:\Users\roger\Product Search - CleverDeals\Docapole\Dec 2025_march 2026 Stocktake.xlsx"
DEFAULT_SHEET = "Stockcount List JUNE 2026"


def norm(s):
    if s is None:
        return None
    # Excel free-text often has stray newlines / doubled spaces; the encoding
    # also mangles the pound sign, but pack-sizes only need plain text.
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def slugify(*parts):
    base = "-".join(p for p in parts if p)
    base = base.lower()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    return base


def build(src_path, sheet_name):
    wb = openpyxl.load_workbook(src_path)
    ws = wb[sheet_name]

    # First find the last row that carries an item/heading in column A.
    last = 0
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value):
            last = r

    # Pre-compute which bold rows are areas (bold row whose next non-empty
    # column-A row is also bold).
    rows = []
    for r in range(2, last + 1):
        v = norm(ws.cell(r, 1).value)
        if not v:
            continue
        rows.append((r, v, bool(ws.cell(r, 1).font.bold)))

    is_area = {}
    for i, (r, v, bold) in enumerate(rows):
        if not bold:
            continue
        nxt = rows[i + 1] if i + 1 < len(rows) else None
        is_area[r] = bool(nxt and nxt[2])

    items = []
    area = None
    section = None
    seen_keys = {}
    order = 0
    for r, v, bold in rows:
        if bold:
            if is_area.get(r):
                area = v
                section = None
            else:
                section = v
            continue
        order += 1
        pack = norm(ws.cell(r, 2).value)
        supplier = norm(ws.cell(r, 3).value)
        key = slugify(section or area or "general", v)
        # Guard against duplicate item names within a section (e.g. a colour
        # repeated) so keys stay stable + unique.
        if key in seen_keys:
            seen_keys[key] += 1
            key = f"{key}-{seen_keys[key]}"
        else:
            seen_keys[key] = 1
        items.append({
            "key": key,
            "area": area,
            "section": section,
            "name": v,
            "pack": pack,
            "supplier": supplier,
            "order": order,
        })

    return items


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    sheet = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SHEET
    items = build(src, sheet)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "src", "data")
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "catalogue.json")

    payload = {
        "source": os.path.basename(src),
        "sheet": sheet,
        "itemCount": len(items),
        "items": items,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    areas = []
    for it in items:
        a = it["area"]
        if a not in areas:
            areas.append(a)
    print(f"[build-catalogue] wrote {len(items)} items to {out_path}")
    print(f"[build-catalogue] areas: {', '.join(str(a) for a in areas)}")


if __name__ == "__main__":
    main()
